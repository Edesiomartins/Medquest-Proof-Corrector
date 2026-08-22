"""Template e importação de provas via planilha Excel (.xlsx)."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

_TEMPLATE_QUESTION_ROWS = 12
_METADATA_ROWS = (
    ("titulo", "Título da prova", 3),
    ("disciplina", "Disciplina", 4),
    ("curso", "Curso", 5),
    ("turma", "Turma", 6),
    ("valor_padrao", "Valor padrão por questão", 7),
)
_QUESTION_HEADER_ROW = 9
_QUESTION_FIRST_DATA_ROW = 10


def build_exam_xlsx_template(*, is_practical: bool = False) -> bytes:
    """Gera o arquivo .xlsx modelo para importação de prova."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Prova"

    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="047857")
    label_font = Font(bold=True)

    exam_kind = "prática" if is_practical else "discursiva"
    ws["A1"] = f"Template de prova {exam_kind} — medquestcorrector"
    ws["A1"].font = title_font
    ws.merge_cells("A1:E1")

    ws["A2"] = (
        "Preencha os dados gerais (coluna B) e as questões abaixo. "
        "Linhas sem enunciado são ignoradas na importação."
    )
    ws.merge_cells("A2:E2")
    ws["A2"].alignment = Alignment(wrap_text=True)

    for _key, label, row in _METADATA_ROWS:
        ws.cell(row=row, column=1, value=label).font = label_font
        if _key == "valor_padrao":
            ws.cell(row=row, column=2, value=1.0)

    headers = ("Nº", "Enunciado", "Resposta esperada", "Critérios de correção", "Pontuação")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=_QUESTION_HEADER_ROW, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    example_enunciado = (
        "Identifique a estrutura anatômica indicada e descreva sua função."
        if is_practical
        else "Descreva o mecanismo fisiológico solicitado no enunciado."
    )
    example_answer = (
        "Ex.: músculo bíceps braquial — flexão do antebraço sobre o braço."
        if is_practical
        else "Ex.: resposta objetiva com os conceitos essenciais esperados."
    )
    example_criteria = "Ex.: 0,5 pt por conceito correto; 0,5 pt por exemplificação."

    for idx in range(_TEMPLATE_QUESTION_ROWS):
        row = _QUESTION_FIRST_DATA_ROW + idx
        ws.cell(row=row, column=1, value=idx + 1)
        if idx == 0:
            ws.cell(row=row, column=2, value=example_enunciado)
            ws.cell(row=row, column=3, value=example_answer)
            ws.cell(row=row, column=4, value=example_criteria)
            ws.cell(row=row, column=5, value=1.0)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 12

    for row in range(_QUESTION_FIRST_DATA_ROW, _QUESTION_FIRST_DATA_ROW + _TEMPLATE_QUESTION_ROWS):
        for col in (2, 3, 4):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def parse_exam_xlsx(raw: bytes) -> dict:
    """
    Lê a planilha preenchida e devolve o mesmo formato de `_parse_discursive_docx`:
    `{metadata, questions, warnings}`.
    """
    wb = load_workbook(BytesIO(raw), data_only=True)
    ws = wb["Prova"] if "Prova" in wb.sheetnames else wb.active

    metadata: dict[str, str] = {}
    for key, label, row in _METADATA_ROWS:
        cell_val = ws.cell(row=row, column=2).value
        if cell_val is None:
            continue
        text = str(cell_val).strip()
        if not text:
            continue
        if key == "valor_padrao":
            metadata[key] = text.replace(",", ".")
        else:
            metadata[key] = text

    default_score = _parse_score(metadata.get("valor_padrao", "1"), 1.0)
    warnings: list[str] = []
    questions: list[dict] = []
    qnum = 0

    for row in range(_QUESTION_FIRST_DATA_ROW, ws.max_row + 1):
        enunciado = _cell_text(ws, row, 2)
        if not enunciado:
            continue

        raw_num = ws.cell(row=row, column=1).value
        if raw_num is not None and str(raw_num).strip():
            try:
                qnum = int(float(str(raw_num).replace(",", ".")))
            except ValueError:
                qnum += 1
                warnings.append(f"Linha {row}: número inválido; usado {qnum}.")
        else:
            qnum += 1

        resposta = _cell_text(ws, row, 3)
        criterios = _cell_text(ws, row, 4)
        score_raw = ws.cell(row=row, column=5).value
        max_score = (
            _parse_score(str(score_raw), default_score)
            if score_raw is not None and str(score_raw).strip()
            else default_score
        )

        if not resposta:
            warnings.append(f"Questão {qnum} (linha {row}): resposta esperada não informada.")

        questions.append(
            {
                "question_number": qnum,
                "question_text": enunciado,
                "expected_answer": resposta or "Resposta esperada não informada.",
                "correction_criteria": criterios,
                "max_score": max_score,
            }
        )

    return {"metadata": metadata, "questions": questions, "warnings": warnings}


def _cell_text(ws, row: int, col: int) -> str:
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()


def _parse_score(value: str, default: float) -> float:
    try:
        return float(str(value).replace(",", ".").strip())
    except (TypeError, ValueError):
        return default
