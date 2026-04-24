"""Exporta resultados de correção para planilha Excel."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def export_results_xlsx(
    exam_name: str,
    questions: list[dict],
    results: list[dict],
) -> bytes:
    """
    Gera um arquivo .xlsx com as notas.

    questions: [{"number": 1, "text": "...", "max_score": 1.0}, ...]
    results: [{
        "student_name": str,
        "registration_number": str,
        "curso": str,
        "turma": str,
        "scores": {1: 0.75, 2: 1.0, ...},  # question_number -> final_score
        "total": float,
    }, ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Notas"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Matrícula", "Nome", "Curso", "Turma"]
    for q in questions:
        headers.append(f"Q{q['number']} ({q['max_score']})")
    headers.append("TOTAL")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["registration_number"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=r["student_name"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=r.get("curso", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value=r.get("turma", "")).border = thin_border

        for q_idx, q in enumerate(questions):
            score = r["scores"].get(q["number"])
            cell = ws.cell(row=row_idx, column=5 + q_idx, value=score if score is not None else "—")
            cell.alignment = center
            cell.border = thin_border

        total_cell = ws.cell(row=row_idx, column=5 + len(questions), value=r["total"])
        total_cell.font = Font(bold=True)
        total_cell.alignment = center
        total_cell.border = thin_border

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14

    # Título acima
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=exam_name)
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
