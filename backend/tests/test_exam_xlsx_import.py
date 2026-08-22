from io import BytesIO

from openpyxl import load_workbook

from app.services.exam_import.exam_xlsx import build_exam_xlsx_template, parse_exam_xlsx


def test_build_exam_xlsx_template_has_metadata_and_question_headers():
    raw = build_exam_xlsx_template(is_practical=False)
    wb = load_workbook(BytesIO(raw))
    ws = wb["Prova"]

    assert ws["A3"].value == "Título da prova"
    assert ws["A9"].value == "Nº"
    assert ws["B9"].value == "Enunciado"
    assert ws["B10"].value  # linha de exemplo


def test_parse_exam_xlsx_reads_metadata_and_questions():
    raw = build_exam_xlsx_template(is_practical=False)
    wb = load_workbook(BytesIO(raw))
    ws = wb["Prova"]
    ws["B3"] = "Farmacologia I"
    ws["B6"] = "Turma A"
    ws["B10"] = "Enunciado questão 1"
    ws["C10"] = "Resposta 1"
    ws["D10"] = "Critério 1"
    ws["B11"] = "Enunciado questão 2"
    ws["C11"] = "Resposta 2"
    buf = BytesIO()
    wb.save(buf)

    parsed = parse_exam_xlsx(buf.getvalue())

    assert parsed["metadata"]["titulo"] == "Farmacologia I"
    assert parsed["metadata"]["turma"] == "Turma A"
    assert len(parsed["questions"]) == 2
    assert parsed["questions"][0]["question_text"] == "Enunciado questão 1"
    assert parsed["questions"][1]["expected_answer"] == "Resposta 2"


def test_parse_exam_xlsx_ignores_empty_question_rows():
    raw = build_exam_xlsx_template(is_practical=True)
    wb = load_workbook(BytesIO(raw))
    ws = wb["Prova"]
    ws["B3"] = "Prova prática"
    ws["B10"] = "Única questão"
    ws["C10"] = "Resposta"
    buf = BytesIO()
    wb.save(buf)

    parsed = parse_exam_xlsx(buf.getvalue())

    assert len(parsed["questions"]) == 1
    assert parsed["questions"][0]["question_number"] == 1
